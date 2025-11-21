from django.shortcuts import render , redirect, HttpResponse
from django.contrib.auth import authenticate, login as auth_login , logout 
from django.contrib import messages
from django.contrib.auth.models import User
from .models import UserProfile  ,EnergyType
from django.contrib.auth.decorators import login_required

from django.contrib import messages
from django.shortcuts import redirect

@login_required
def register_user(request):
    if request.method == 'POST':
        username = request.POST.get("username")
        email = request.POST.get("email")
        password = request.POST.get("password")
        is_superuser = request.POST.get("is_superuser", 'off') == 'on'
        is_staff = request.POST.get("is_staff", 'off') == 'on'
        is_active = request.POST.get("is_active", 'off') == 'on'
        selected_ids = request.POST.getlist("energy_types")

        user = User(
            username=username,
            email=email,
            is_superuser=is_superuser,
            is_staff=is_staff,
            is_active=is_active
        )
        user.set_password(password)
        user.save()

        profile = UserProfile.objects.create(user=user)
        profile.energy_types.set(selected_ids)

        return redirect('user_list')  # or wherever you want to show the modal

    else:
        energy_choices = EnergyType.objects.all()
    return render(request, 'signup.html', {
    "energy_choices": EnergyType.objects.all(),
    "user_registered": True
    })

def login_view(request):
    if request.method == 'POST':
        username = request.POST.get('username')
        password = request.POST.get('password')
        user = authenticate(request, username=username, password=password)
        if user is not None:
            auth_login(request, user)
            return redirect('/dashboard') 
        else:
            messages.error(request, "Invalid username or password.")
    return render(request, "login.html")


def logout_view(request):
    logout(request)
    messages.success(request, "You have been logged out.")
    return redirect('login')

from django.core.paginator import Paginator
from django.contrib.auth.models import User
from accounts.models import UserProfile

from django.db.models import Q

@login_required
def user_list(request):
    username = request.GET.get('username', '')
    energy_type = request.GET.get('energy_type', '')
    status = request.GET.get('status', '')

    users = User.objects.all().select_related('userprofile').distinct()

    if username:
        users = users.filter(Q(username__icontains=username) | Q(email__icontains=username))

    if energy_type:
        users = users.filter(userprofile__energy_types__name__iexact=energy_type)

    if status == 'active':
        users = users.filter(is_active=True)
    elif status == 'inactive':
        users = users.filter(is_active=False)

    paginator = Paginator(users, 10)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    return render(request, 'user_list.html', {
        'users': page_obj,
        'page_obj': page_obj,
        'username': username,
        'energy_type': energy_type,
        'status': status,
    })

from django.shortcuts import get_object_or_404, redirect
from django.contrib.auth.models import User
from accounts.models import UserProfile

@login_required
def delete_user(request, user_id):
    user = get_object_or_404(User, id=user_id)
    if request.method == 'POST':
        user.delete()
        return redirect('user_list')  # Assuming you named the URL
    return HttpResponse("Invalid request.")

from django.contrib.auth.models import update_last_login
from django.contrib.auth import get_user_model
from django.contrib.auth.signals import user_logged_in
from django.contrib.admin.models import LogEntry
from django.utils import timezone  # ✅ Use only this
 

@login_required
def edit_user(request, user_id):
    user = get_object_or_404(User, id=user_id)
    profile, _ = UserProfile.objects.get_or_create(user=user)

    if request.method == 'POST':
        user.username = request.POST.get("username")
        user.email = request.POST.get("email")
        user.is_active = 'is_active' in request.POST
        user.is_staff = 'is_staff' in request.POST
        user.is_superuser = 'is_superuser' in request.POST

        new_password = request.POST.get("password")
        if new_password:
            user.set_password(new_password)
            profile.password_updated_at = timezone.now()  # ✅ Update timestamp

        user.save()

        selected_energy_ids = request.POST.getlist("energy_types")
        profile.energy_types.set(selected_energy_ids)
        profile.save()

        messages.success(request, "User updated successfully.")
        return redirect('user_list')

    # GET
    energy_choices = EnergyType.objects.all()
    selected_energy_ids = profile.energy_types.values_list('id', flat=True)

    return render(request, 'edit_user.html', {
        'user_obj': user,
        'energy_choices': energy_choices,
        'selected_energy_ids': selected_energy_ids,
        'current_login': request.user.last_login,
        'last_login': user.last_login,
        'created_on': user.date_joined,
        'last_password_change': profile.password_updated_at  # ✅ Correct field name
    })


from django.shortcuts import render, redirect
from .models import EnergyType
from django.contrib import messages
from django.contrib.auth.decorators import login_required

from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.shortcuts import render, redirect
from .models import EnergyType

@login_required
def add_energy_type(request):
    if request.method == 'POST':
        if 'delete_id' in request.POST:
            # 🗑️ Handle energy type deletion
            delete_id = request.POST.get('delete_id')
            try:
                EnergyType.objects.get(id=delete_id).delete()
                messages.success(request, "Energy type deleted successfully!")
            except EnergyType.DoesNotExist:
                messages.error(request, "Energy type not found.")
            return redirect('add_energy_type')  # 👈 stay on same page

        # ➕ Handle energy type addition
        name = request.POST.get('name', '').strip()
        if name:
            obj, created = EnergyType.objects.get_or_create(name=name)
            if created:
                messages.success(request, "Energy type added successfully!")
            else:
                messages.warning(request, "Energy type already exists.")
        else:
            messages.error(request, "Please enter a valid energy type name.")
        return redirect('add_energy_type')  # 👈 stay on same page

    # 📋 Show existing energy types
    energy_types = EnergyType.objects.all().order_by('name')
    return render(request, 'add_energy_type.html', {'energy_types': energy_types})

import pandas as pd
from django.db import connection
from .models import Provider
from django.shortcuts import render, redirect
from django.contrib import messages
from django.contrib.auth.decorators import login_required

import os
import pandas as pd
from django.shortcuts import render, redirect
from django.contrib import messages
from django.db import connection, IntegrityError
from django.core.files.storage import FileSystemStorage
from django.contrib.auth.decorators import login_required
from .models import Provider

import os
import pandas as pd
from django.shortcuts import render, redirect
from django.contrib import messages
from django.db import connection, IntegrityError
from django.core.files.storage import FileSystemStorage
from django.contrib.auth.decorators import login_required
from .models import Provider

from accounts.models import EnergyType  # ✅ Import EnergyType


from django.contrib.auth.decorators import login_required
from django.shortcuts import render, redirect
from django.contrib import messages
from django.core.files.storage import FileSystemStorage
from django.db import connection
import pandas as pd

from .models import Provider
from accounts.models import EnergyType


from django.contrib.auth.decorators import login_required
from django.shortcuts import render, redirect
from django.contrib import messages
from django.core.files.storage import FileSystemStorage
from django.db import connection
from django.contrib.auth.models import User
import pandas as pd
from accounts.models import Provider
from accounts.models import EnergyType
from django.shortcuts import render, redirect
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.core.files.storage import FileSystemStorage
from django.db import connection
from django.contrib.auth.models import User
from accounts.models import Provider, EnergyType

import pandas as pd
import os
import re
import xlrd
from openpyxl import Workbook

from django.contrib.auth.decorators import login_required
from django.shortcuts import render, redirect
from django.contrib import messages
from django.core.files.storage import FileSystemStorage
from django.db import connection
from django.contrib.auth.models import User
import pandas as pd
import os
import re
import xlrd
from openpyxl import Workbook
from accounts.models import UserProvider 
from accounts.models import Provider, EnergyType


# --- helper: convert old .xls to .xlsx ---
def convert_xls_to_xlsx(xls_path, xlsx_path):
    """Convert old .xls file to .xlsx using xlrd + openpyxl"""
    book = xlrd.open_workbook(xls_path)
    sheet = book.sheet_by_index(0)
    wb = Workbook()
    ws = wb.active

    for row in range(sheet.nrows):
        ws.append(sheet.row_values(row))

    wb.save(xlsx_path)
    return xlsx_path


@login_required
def add_provider_with_structure(request):
    energy_types = EnergyType.objects.all()
    staff_users = User.objects.filter(is_superuser=False)

    # ------------------ POST logic ------------------
    if request.method == 'POST':

        # ✅ DELETE specific table
        if 'delete_table_name' in request.POST:
            table_to_delete = request.POST.get('delete_table_name')
            try:
                with connection.cursor() as cursor:
                    cursor.execute(f"DROP TABLE IF EXISTS `{table_to_delete}`")
                messages.success(request, f"Table '{table_to_delete}' deleted successfully.")
            except Exception as e:
                messages.error(request, f"Error deleting table: {str(e)}")
            return redirect("add_provider")

        # ✅ DELETE provider and related tables
        elif 'delete_id' in request.POST:
            delete_id = request.POST.get('delete_id')
            try:
                provider = Provider.objects.get(id=delete_id)
                provider_clean = provider.name.lower().replace(' ', '_')

                with connection.cursor() as cursor:
                    for energy in EnergyType.objects.all():
                        for user in staff_users:
                            table_name = f"{user.username.lower()}_{provider_clean}_{energy.name.lower().replace(' ', '_')}"
                            table_break = f"{table_name}_breakdowndata"
                            cursor.execute(f"DROP TABLE IF EXISTS `{table_name}`")
                            cursor.execute(f"DROP TABLE IF EXISTS `{table_break}`")

                provider.delete()
                messages.success(request, "Provider and all related tables deleted successfully.")
            except Exception as e:
                messages.error(request, f"Error deleting provider: {str(e)}")
            return redirect("add_provider")

        # ✅ ADD new provider and create structure
        provider_name = request.POST.get("provider_name", "").strip().lower().replace(' ', '_')
        energy_type_id = request.POST.get("energy_type")
        structure_file = request.FILES.get("structure_file")
        selected_username = request.POST.get("selected_user")

        if not provider_name or not energy_type_id or not structure_file or not selected_username:
            messages.error(request, "All fields are required: provider, energy type, file, and user.")
            return redirect("add_provider")

        # ✅ Fetch related objects
        try:
            user_obj = User.objects.get(username=selected_username)
        except User.DoesNotExist:
            messages.error(request, "Invalid user selected.")
            return redirect("add_provider")

        try:
            energy_type = EnergyType.objects.get(id=energy_type_id)
            energy_type_name = energy_type.name.strip().lower().replace(' ', '_')
        except EnergyType.DoesNotExist:
            messages.error(request, "Invalid energy type selected.")
            return redirect("add_provider")

        provider_obj, created = Provider.objects.get_or_create(name=provider_name)
        UserProvider.objects.get_or_create(user=user_obj, provider=provider_obj)


        # ✅ Save uploaded file temporarily
        fs = FileSystemStorage()
        filename = fs.save(structure_file.name, structure_file)
        file_path = fs.path(filename)

        try:
            # ---------- Detect file extension ----------
            file_ext = os.path.splitext(filename)[1].lower()

            if file_ext == ".csv":
                sheets = {"main": pd.read_csv(file_path)}

            elif file_ext == ".xls":
                try:
                    xlsx_path = file_path + "x"
                    convert_xls_to_xlsx(file_path, xlsx_path)
                    sheets = pd.read_excel(xlsx_path, engine="openpyxl", sheet_name=None)
                    os.remove(xlsx_path)
                except Exception:
                    df_list = pd.read_html(file_path)
                    sheets = {"main": df_list[0]}

            elif file_ext in [".xlsx", ".xlsm", ".xlsb"]:
                sheets = pd.read_excel(file_path, engine="openpyxl", sheet_name=None)

            elif file_ext == ".ods":
                sheets = pd.read_excel(file_path, engine="odf", sheet_name=None)

            else:
                messages.error(request, f"Unsupported file type: {file_ext}")
                return redirect("add_provider")

            created_tables = []

            # ---------- Iterate through sheets ----------
            for sheet_name, df in sheets.items():
                # skip empty sheets
                if df.empty:
                    continue

                sheet_clean = sheet_name.strip().lower()

                # ✅ Identify generation vs breakdown
                if "breakdown" in sheet_clean:
                    table_name = f"{selected_username.lower()}_{provider_name}_{energy_type_name}_breakdowndata"
                else:
                    table_name = f"{selected_username.lower()}_{provider_name}_{energy_type_name}"

                # ✅ Clean column names
                df.columns = [
                    re.sub(r'\W+', '_', str(col).strip()).lower().strip('_')
                    for col in df.columns
                ]

                # ✅ Build SQL table
                column_defs = [f"`{col}` TEXT" for col in df.columns]
                column_defs += ["`provider` TEXT", "`energy_type` TEXT", "`uploaded_by` TEXT"]

                create_sql = f"""
                    CREATE TABLE IF NOT EXISTS `{table_name}` (
                        `id` INT AUTO_INCREMENT PRIMARY KEY,
                        {", ".join(column_defs)}
                    );
                """

                with connection.cursor() as cursor:
                    cursor.execute(create_sql)

                created_tables.append(table_name)

            if created_tables:
                messages.success(
                    request,
                    f"✅ Created table(s): {', '.join(created_tables)} for provider '{provider_name}'."
                )
            else:
                messages.warning(request, "No valid sheets found to create tables.")

        except Exception as e:
            messages.error(request, f"❌ Error creating table: {str(e)}")

        finally:
            fs.delete(filename)

        return redirect("add_provider")

    # ------------------ GET logic ------------------
    providers = Provider.objects.all()

    with connection.cursor() as cursor:
        cursor.execute("SHOW TABLES;")
        all_tables = [row[0] for row in cursor.fetchall()]

    provider_table_map = {}
    for provider in providers:
        key = provider.name.strip().lower().replace(' ', '_')
        related_tables = [t for t in all_tables if f"_{key}_" in t]
        provider_table_map[provider.name] = related_tables

    return render(request, 'add_provider.html', {
        'providers': providers,
        'energy_types': energy_types,
        'provider_table_map': provider_table_map,
        'users': staff_users,
    })


from django.contrib.auth.decorators import login_required
from django.shortcuts import render
from django.db import connection
from accounts.models import Provider

from django.contrib.auth.decorators import login_required
from django.shortcuts import render, redirect
from django.db import connection
from accounts.models import Provider
from django.contrib import messages

from django.contrib import messages
from django.shortcuts import redirect
from django.db import connection
from accounts.models import Provider

from django.shortcuts import render, redirect
from django.contrib import messages
from django.db import connection
from .models import Provider  # or your actual Provider model

from django.shortcuts import render, redirect
from django.contrib import messages
from django.db import connection
from .models import Provider  # ✅ Moved this to the top — keep it here

def view_existing_providers(request):
    if request.method == "POST":
        table_to_delete = request.POST.get("delete_table_name")
        if table_to_delete:
            try:
                with connection.cursor() as cursor:
                    cursor.execute(f"DROP TABLE IF EXISTS `{table_to_delete}`")
                messages.success(request, f"Table {table_to_delete} deleted successfully.")
            except Exception as e:
                messages.error(request, f"❌ Error deleting table: {str(e)}")
            return redirect("existing_providers")

        provider_id = request.POST.get("delete_provider_id")
        if provider_id:
            try:
                provider = Provider.objects.get(id=provider_id)
                with connection.cursor() as cursor:
                    cursor.execute("SHOW TABLES;")
                    all_tables = [row[0] for row in cursor.fetchall()]
                    provider_tables = [t for t in all_tables if provider.name.lower() in t.lower()]
                    for table in provider_tables:
                        cursor.execute(f"DROP TABLE IF EXISTS `{table}`")
                provider.delete()
                messages.success(request, f"✅ Provider '{provider.name}' and all related tables deleted.")
            except Provider.DoesNotExist:
                messages.error(request, "❌ Provider not found.")
            except Exception as e:
                messages.error(request, f"❌ Error deleting provider: {str(e)}")
            return redirect("existing_providers")

    # ✅ GET request — display providers and their tables
    providers = Provider.objects.all()

    with connection.cursor() as cursor:
        cursor.execute("SHOW TABLES;")
        all_tables = [row[0] for row in cursor.fetchall()]

    provider_table_map = {}
    for provider in providers:
        matching_tables = [t for t in all_tables if provider.name.lower() in t.lower()]
        provider_table_map[provider.name] = matching_tables

    return render(request, "view_existing_providers.html", {
        "providers": providers,
        "provider_table_map": provider_table_map
    })










from django.contrib.auth.decorators import login_required
from django.shortcuts import render, redirect
from django.contrib import messages
from django.core.files.storage import FileSystemStorage
from django.db import connection
from django.contrib.auth.models import User
import pandas as pd
import os
import re
import xlrd
from openpyxl import Workbook
from accounts.models import EnergyType


from django.contrib.auth.decorators import login_required
from django.shortcuts import render, redirect
from django.contrib import messages
from django.core.files.storage import FileSystemStorage
from django.db import connection
import pandas as pd
import os, re, xlrd
from openpyxl import Workbook


# One common DSM table
DSM_TABLE_NAME = "dsm_data"


# --- helper: convert .xls to .xlsx ---
def convert_xls_to_xlsx(xls_path, xlsx_path):
    book = xlrd.open_workbook(xls_path)
    sheet = book.sheet_by_index(0)
    wb = Workbook()
    ws = wb.active
    for row in range(sheet.nrows):
        ws.append(sheet.row_values(row))
    wb.save(xlsx_path)
    return xlsx_path


@login_required
def add_dsm_structure(request):

    # ----------- DELETE DSM TABLE -----------
    if request.method == "POST" and "delete_dsm_table" in request.POST:
        try:
            with connection.cursor() as cursor:
                cursor.execute(f"DROP TABLE IF EXISTS `{DSM_TABLE_NAME}`")
            messages.success(request, f"🗑️ Deleted table '{DSM_TABLE_NAME}' successfully.")
        except Exception as e:
            messages.error(request, f"❌ Error deleting table: {e}")
        return redirect("add_dsm_structure")

    # ----------- ADD SINGLE COLUMN -----------
    if request.method == "POST" and "add_column" in request.POST:

        col_name = request.POST.get("column_name", "").strip()
        col_type = request.POST.get("column_type", "TEXT").strip().upper()

        if not col_name:
            messages.error(request, "Please enter a column name.")
            return redirect("add_dsm_structure")

        col_name = re.sub(r"\W+", "_", col_name.lower()).strip("_")
        if not col_name:
            messages.error(request, "Invalid column name.")
            return redirect("add_dsm_structure")

        valid_types = ["TEXT", "INT", "FLOAT", "DATE"]
        if col_type not in valid_types:
            col_type = "TEXT"

        try:
            with connection.cursor() as cursor:
                cursor.execute("SHOW TABLES;")
                tables = [r[0] for r in cursor.fetchall()]
                if DSM_TABLE_NAME not in tables:
                    messages.error(request, "⚠️ DSM table does not exist. Upload structure file first.")
                    return redirect("add_dsm_structure")

                cursor.execute(f"SHOW COLUMNS FROM `{DSM_TABLE_NAME}`;")
                existing = [r[0].lower() for r in cursor.fetchall()]
                if col_name in existing:
                    messages.warning(request, f"Column '{col_name}' already exists.")
                    return redirect("add_dsm_structure")

                cursor.execute(f"ALTER TABLE `{DSM_TABLE_NAME}` ADD COLUMN `{col_name}` {col_type};")

            messages.success(request, f"✅ Column '{col_name}' ({col_type}) added successfully!")
        except Exception as e:
            messages.error(request, f"❌ Failed to add column: {e}")

        return redirect("add_dsm_structure")

    # ----------- UPLOAD STRUCTURE FILE (DROP + RECREATE) -----------
    if request.method == "POST" and "structure_file" in request.FILES:

        structure_file = request.FILES["structure_file"]
        fs = FileSystemStorage()
        filename = fs.save(structure_file.name, structure_file)
        file_path = fs.path(filename)

        try:
            ext = os.path.splitext(filename)[1].lower()

            if ext == ".csv":
                df = pd.read_csv(file_path)
            elif ext == ".xls":
                xlsx_path = file_path + "x"
                convert_xls_to_xlsx(file_path, xlsx_path)
                df = pd.read_excel(xlsx_path, engine="openpyxl")
                os.remove(xlsx_path)
            elif ext in [".xlsx", ".xlsm", ".xlsb"]:
                df = pd.read_excel(file_path, engine="openpyxl")
            elif ext == ".ods":
                df = pd.read_excel(file_path, engine="odf")
            else:
                messages.error(request, f"Unsupported file type: {ext}")
                return redirect("add_dsm_structure")

            if df.empty:
                messages.warning(request, "Uploaded file is empty.")
                return redirect("add_dsm_structure")

            # Clean column names
            cleaned = []
            seen = {}

            for col in df.columns:
                c = re.sub(r"\W+", "_", str(col).strip()).lower().strip("_") or "col"
                if c not in seen:
                    seen[c] = 0
                    cleaned.append(c)
                else:
                    seen[c] += 1
                    cleaned.append(f"{c}_{seen[c]}")

            # Extra internal columns
            extra_cols = [
                "`username` TEXT",
                "`energy_type` TEXT",
                "`provider` TEXT",
                
            ]

            dynamic = [f"`{c}` TEXT" for c in cleaned]

            create_sql = f"""
                CREATE TABLE IF NOT EXISTS `{DSM_TABLE_NAME}` (
                    `id` INT AUTO_INCREMENT PRIMARY KEY,
                    {', '.join(dynamic + extra_cols)}
                );
            """

            with connection.cursor() as cursor:
                cursor.execute(f"DROP TABLE IF EXISTS `{DSM_TABLE_NAME}`")
                cursor.execute(create_sql)

            messages.success(request, f"🚀 DSM Table '{DSM_TABLE_NAME}' created with {len(cleaned)} columns!")

        except Exception as e:
            messages.error(request, f"❌ Failed to create table: {e}")
        finally:
            fs.delete(filename)

        return redirect("add_dsm_structure")

    # ----------- CHECK TABLE -----------
    with connection.cursor() as cursor:
        cursor.execute("SHOW TABLES;")
        tables = [row[0] for row in cursor.fetchall()]
    table_exists = DSM_TABLE_NAME in tables

    return render(request, "add_dsm_structure.html", {
        "dsm_table_name": DSM_TABLE_NAME,
        "table_exists": table_exists,
    })




from django.contrib.auth.decorators import login_required
from django.shortcuts import render, redirect
from django.contrib import messages
from django.core.files.storage import FileSystemStorage
from django.db import connection
import pandas as pd
import os
import re
import xlrd
from openpyxl import Workbook

# ---------- constants ----------
PM_TABLE_NAME = "preventive_maintenance_data"

from django.contrib.auth.decorators import login_required
from django.shortcuts import render, redirect
from django.contrib import messages
from django.core.files.storage import FileSystemStorage
from django.db import connection
import pandas as pd
import os, re, xlrd
from openpyxl import Workbook

PM_TABLE_NAME = "preventive_maintenance_data"

# --- helper: convert old .xls to .xlsx ---
def convert_xls_to_xlsx(xls_path, xlsx_path):
    book = xlrd.open_workbook(xls_path)
    sheet = book.sheet_by_index(0)
    wb = Workbook()
    ws = wb.active
    for row in range(sheet.nrows):
        ws.append(sheet.row_values(row))
    wb.save(xlsx_path)
    return xlsx_path


@login_required
def add_pm_structure(request):
    """
    Upload or modify preventive maintenance table structure.
    - Upload file → drop & recreate structure.
    - Manual form → add a new column without replacing table.
    """

    # --- Delete Table ---
    if request.method == "POST" and "delete_pm_table" in request.POST:
        try:
            with connection.cursor() as cursor:
                cursor.execute(f"DROP TABLE IF EXISTS `{PM_TABLE_NAME}`")
            messages.success(request, f"✅ Table '{PM_TABLE_NAME}' deleted successfully.")
        except Exception as e:
            messages.error(request, f"❌ Error deleting table: {str(e)}")
        return redirect("add_pm_structure")

    # --- Add Single Column Manually ---
    if request.method == "POST" and "add_column" in request.POST:
        col_name = request.POST.get("column_name", "").strip()
        col_type = request.POST.get("column_type", "TEXT").strip().upper()

        if not col_name:
            messages.error(request, "Please provide a column name.")
            return redirect("add_pm_structure")

        # clean column name
        col_name = re.sub(r"\W+", "_", col_name.lower()).strip("_")
        if not col_name:
            messages.error(request, "Invalid column name.")
            return redirect("add_pm_structure")

        # validate type
        valid_types = ["TEXT", "INT", "FLOAT", "DATE"]
        if col_type not in valid_types:
            col_type = "TEXT"

        try:
            with connection.cursor() as cursor:
                cursor.execute(f"SHOW TABLES;")
                tables = [r[0] for r in cursor.fetchall()]
                if PM_TABLE_NAME not in tables:
                    messages.error(request, "⚠️ Table does not exist yet. Create structure first.")
                    return redirect("add_pm_structure")

                # check if already exists
                cursor.execute(f"SHOW COLUMNS FROM `{PM_TABLE_NAME}`;")
                existing = [r[0].lower() for r in cursor.fetchall()]
                if col_name.lower() in existing:
                    messages.warning(request, f"Column '{col_name}' already exists.")
                    return redirect("add_pm_structure")

                cursor.execute(f"ALTER TABLE `{PM_TABLE_NAME}` ADD COLUMN `{col_name}` {col_type};")
            messages.success(request, f"✅ Column '{col_name}' ({col_type}) added successfully!")
        except Exception as e:
            messages.error(request, f"❌ Failed to add column: {e}")
        return redirect("add_pm_structure")

    # --- Upload structure file (replace table) ---
    if request.method == "POST" and "structure_file" in request.FILES:
        structure_file = request.FILES.get("structure_file")
        if not structure_file:
            messages.error(request, "Please choose a structure file.")
            return redirect("add_pm_structure")

        fs = FileSystemStorage()
        filename = fs.save(structure_file.name, structure_file)
        file_path = fs.path(filename)

        try:
            ext = os.path.splitext(filename)[1].lower()
            if ext == ".csv":
                df = pd.read_csv(file_path)
            elif ext == ".xls":
                xlsx_path = file_path + "x"
                convert_xls_to_xlsx(file_path, xlsx_path)
                df = pd.read_excel(xlsx_path, engine="openpyxl", sheet_name=0)
                os.remove(xlsx_path)
            elif ext in [".xlsx", ".xlsm", ".xlsb"]:
                df = pd.read_excel(file_path, engine="openpyxl", sheet_name=0)
            elif ext == ".ods":
                df = pd.read_excel(file_path, engine="odf", sheet_name=0)
            else:
                messages.error(request, f"Unsupported file type: {ext}")
                return redirect("add_pm_structure")

            if df.empty:
                messages.warning(request, "Uploaded file has no data.")
                return redirect("add_pm_structure")

            # clean column names
            cleaned = []
            seen = {}
            for col in df.columns:
                c = re.sub(r"\W+", "_", str(col).strip()).lower().strip("_") or "col"
                if c not in seen:
                    seen[c] = 0
                    cleaned.append(c)
                else:
                    seen[c] += 1
                    cleaned.append(f"{c}_{seen[c]}")

            # build SQL
            extra_cols = [
                "`username` TEXT",
                "`energy_type` TEXT",
                "`checkpoints_period` TEXT",
                "`mw` TEXT",
                "`description` TEXT",
            ]
            dynamic = [f"`{c}` TEXT" for c in cleaned]
            create_sql = f"""
                CREATE TABLE IF NOT EXISTS `{PM_TABLE_NAME}` (
                    id INT AUTO_INCREMENT PRIMARY KEY,
                    {', '.join(dynamic + extra_cols)}
                );
            """

            with connection.cursor() as cursor:
                cursor.execute(f"DROP TABLE IF EXISTS `{PM_TABLE_NAME}`")
                cursor.execute(create_sql)

            messages.success(request, f"✅ Created '{PM_TABLE_NAME}' with {len(cleaned)} columns.")
        except Exception as e:
            messages.error(request, f"❌ Error creating table: {e}")
        finally:
            fs.delete(filename)
        return redirect("add_pm_structure")

    # --- Check table existence ---
    with connection.cursor() as cursor:
        cursor.execute("SHOW TABLES;")
        tables = [r[0] for r in cursor.fetchall()]
    table_exists = PM_TABLE_NAME in tables

    return render(request, "add_pm_structure.html", {
        "pm_table_name": PM_TABLE_NAME,
        "table_exists": table_exists,
    })














